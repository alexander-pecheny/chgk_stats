#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Скрипт для анализа результатов турнира.
Генерирует Excel-файл со статистикой по командам и вопросам.
"""
import argparse
from collections import Counter, defaultdict

import requests
import openpyxl
from openpyxl.styles import Alignment, Font

def get_results(id_):
    """Получает результаты турнира через API."""
    url = f"https://api.rating.chgk.net/tournaments/{id_}/results.json?includeMasksAndControversials=1&includeTeamFlags=1"
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception(f"Ошибка API: {response.status_code} - {response.text}")
    return response.json()

def get_cat(share):
    """Определяет категорию вопроса по доле взявших команд."""
    if 0 <= share <= 0.1:
        return "a"
    if 0.1 < share <= 0.33:
        return "b"
    if 0.33 < share <= 0.66:
        return "c"
    if 0.66 < share <= 0.9:
        return "d"
    if share > 0.9:
        return "e"

def has_required_flag(team_flags, required_flags):
    """Проверяет, есть ли у команды хотя бы один из требуемых флагов."""
    if not required_flags:
        return True  # Если флаги не указаны, учитываем все команды
    
    if not team_flags:
        return False
    
    team_flag_ids = [flag["id"] for flag in team_flags]
    return any(flag_id in team_flag_ids for flag_id in required_flags)

def get_team_flags_string(team_flags, flag_ids=None):
    """Возвращает строку с короткими названиями флагов команды."""
    if not team_flags:
        return ""
    
    if flag_ids:
        # Фильтруем флаги по ID
        filtered_flags = [flag for flag in team_flags if flag["id"] in flag_ids]
    else:
        # Используем все флаги
        filtered_flags = team_flags
    
    return ", ".join(flag["shortName"] for flag in filtered_flags)

def count_non_removed_questions(mask):
    """Подсчитывает количество не снятых вопросов в маске."""
    return sum(1 for v in mask if v != 'X')

def parse_ids(ids_str):
    """Парсит строку с ID, разделенных запятыми."""
    if not ids_str:
        return []
    return [int(id_.strip()) for id_ in ids_str.split(',') if id_.strip().isdigit()]
def main():
    """Основная функция скрипта."""
    parser = argparse.ArgumentParser(description="Анализ результатов турнира")
    parser.add_argument("id", help="ID турнира")
    parser.add_argument("--flags", help="ID флагов команд для фильтрации, разделенные запятыми")
    parser.add_argument("--add_flags", help="ID флагов команд для отображения, разделенные запятыми. Если указано без значения, выводятся все флаги", nargs='?', const='all')
    parser.add_argument("--teams", help="ID команд для отображения их взятий, разделенные запятыми")
    args = parser.parse_args()
    
    # Парсим флаги и ID команд из строк с запятыми
    filter_flag_ids = parse_ids(args.flags) if args.flags else []
    requested_team_ids = parse_ids(args.teams) if args.teams else []
    
    # Парсим флаги для отображения
    if args.add_flags == 'all':
        display_flag_ids = None  # Отображать все флаги
    elif args.add_flags:
        display_flag_ids = parse_ids(args.add_flags)
    else:
        display_flag_ids = []  # Не отображать флаги

    try:
        results = get_results(args.id)
    except Exception as e:
        print(f"Ошибка при получении данных: {e}")
        return
    
    # Проверяем наличие запрошенных команд в турнире
    all_team_ids = {res["team"]["id"] for res in results}
    missing_team_ids = [team_id for team_id in requested_team_ids if team_id not in all_team_ids]
    valid_team_ids = [team_id for team_id in requested_team_ids if team_id in all_team_ids]
    
    # Выводим сообщения о командах, которых нет в турнире
    for team_id in missing_team_ids:
        print(f"Команда {team_id} не участвовала в турнире")
    
    # Если все запрошенные команды отсутствуют, но были запрошены команды
    if requested_team_ids and not valid_team_ids:
        print("Ни одна из указанных команд не участвовала в турнире")
        # Продолжаем работу, но без столбцов для команд
    
    # Фильтрация команд по флагам
    if filter_flag_ids:
        filtered_results = [r for r in results if has_required_flag(r.get("flags", []), filter_flag_ids)]
    else:
        filtered_results = results
    
    # Если после фильтрации не осталось команд, выводим сообщение и завершаем работу
    if not filtered_results:
        print("Не найдено команд с указанными флагами")
        return
    
    # Определяем общее количество вопросов по длине маски первой команды
    if filtered_results and 'mask' in filtered_results[0]:
        total_questions = len(filtered_results[0]['mask'])
    else:
        total_questions = 0
        print("Предупреждение: не удалось определить общее количество вопросов")
    
    # Собираем информацию о командах, которые нужно показать отдельно
    team_masks = {}
    for res in results:
        team_id = res["team"]["id"]
        if team_id in valid_team_ids:
            team_masks[team_id] = list(res["mask"])
    t_rating = Counter()
    t_questions = {}
    t_by_cat = defaultdict(Counter)
    teams_taken = defaultdict(list)
    # Сохраняем информацию о флагах команд
    team_flags_info = {}
    n_teams = len(filtered_results)
    question_data = []  # Данные о вопросах для вывода
    
    # Определяем, какие вопросы были сняты
    removed_questions = set()
    if filtered_results and 'mask' in filtered_results[0]:
        for i, v in enumerate(filtered_results[0]['mask']):
            if v == 'X':
                removed_questions.add(i + 1)  # Номера вопросов начинаются с 1
    
    # Собираем информацию о взятых вопросах и флагах команд
    for res in filtered_results:
        team = (res["team"]["id"], res["current"]["name"])
        t_questions[team] = res["questionsTotal"]
        # Сохраняем флаги команды
        team_flags_info[team] = res.get("flags", [])
        for i, v in enumerate(list(res["mask"])):
            q_num = i + 1
            if v == "1":
                teams_taken[q_num].append(team)
    
    # Обрабатываем все вопросы, включая не взятые и снятые
    all_cats = Counter()
    for q_num in range(1, total_questions + 1):
        teams = teams_taken.get(q_num, [])
        
        # Проверяем, был ли вопрос снят
        is_removed = q_num in removed_questions
        
        # Создаем кортеж с данными о вопросе и взятиях выбранных команд
        question_tuple = [q_num]  # Номер вопроса
        
        if is_removed:
            # Вопрос был снят
            question_tuple.extend([0, "-", "снят"])
        else:
            # Вопрос не был снят
            if q_num in teams_taken:
                # Вопрос взяла хотя бы одна команда
                rating = n_teams - len(teams) + 1
                cat = get_cat(len(teams) / n_teams)
                all_cats[cat] += 1
                for t in teams:
                    t_rating[t] += rating
                    t_by_cat[t][cat] += 1
                taken = len(teams)
                question_tuple.extend([taken, rating, cat])
            else:
                # Вопрос не взяла ни одна команда
                question_tuple.extend([0, n_teams + 1, "гроб"])
        
        # Добавляем информацию о взятиях выбранных команд
        for team_id in valid_team_ids:
            if team_id in team_masks:
                # Индекс в маске на 1 меньше номера вопроса
                mask_value = team_masks[team_id][q_num - 1] if q_num <= len(team_masks[team_id]) else "0"
                question_tuple.append(mask_value)
            else:
                question_tuple.append("-")
        
        question_data.append(tuple(question_tuple))
    
    # Сортируем команды сначала по количеству взятых вопросов, а затем по рейтингу
    sorted_teams = sorted(t_questions, key=lambda x: (t_questions[x], t_rating[x]), reverse=True)
    # Создаем Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Турнир {args.id}"
    
    # Определяем, нужно ли добавлять столбец с флагами
    add_flags_column = args.add_flags is not None
    
    # Задаем заголовки столбцов и их ширину
    header = []
    
    # Добавляем столбец с флагами, если нужно
    if add_flags_column:
        header.append(("Флаги", 15))
    
    # Добавляем основные столбцы
    header.extend([
        ("ID", 7),
        ("Команда", 45),
        ("Рейтинг", 10),
        ("Взято", 10),
        ("Ср. рейт", 10),
        ("A", 3.5),
        ("B", 3.5),
        ("C", 3.5),
        ("D", 3.5),
        ("E", 3.5),
        ("", 3.5),
        ("A", 3.5),
        ("B", 3.5),
        ("C", 3.5),
        ("D", 3.5),
        ("E", 3.5),
        ("", 3.5),  # Пустой столбец для разделения
        ("Вопрос", 10),
        ("Взятия", 10),
        ("Рейтинг", 10),
        ("Категория", 10),
    ])
    
    # Добавляем столбцы для выбранных команд, которые есть в турнире
    for team_id in valid_team_ids:
        header.append((str(team_id), 5))
    
    # Настраиваем заголовки и ширину столбцов
    # Создаем жирный шрифт для заголовков
    bold_font = Font(bold=True)
    
    for i, tup in enumerate(header):
        cell = ws.cell(row=1, column=i + 1)
        cell.value = tup[0]
        ws.column_dimensions[cell.column_letter].width = tup[1]
        
        # Применяем жирный шрифт и выравнивание по центру для всех заголовков
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")
    
    # Заполняем данные о командах
    first = True
    categs = ["a", "b", "c", "d", "e"]
    for t in sorted_teams:
        q = t_questions[t]
        r = t_rating[t]
        # Рассчитываем средний рейтинг
        avg_rating = 0 if q == 0 else round(r / q, 2)
        
        # Создаем строку данных
        row = []
        
        # Добавляем флаги, если нужно
        if add_flags_column:
            row.append(get_team_flags_string(team_flags_info.get(t, []), display_flag_ids))
        
        # Добавляем основные данные
        row.extend([
            t[0],
            t[1],
            r,
            q,
            avg_rating,
        ] + [
            t_by_cat[t][c] for c in categs
        ])
        # Добавляем общую статистику по категориям только в первую строку
        if first:
            row.extend([""] + [all_cats[c] for c in categs])
            first = False
        
        ws.append(row)
    
    # После добавления данных о командах, устанавливаем выравнивание по правому краю для числовых ячеек
    for row_idx in range(2, len(sorted_teams) + 2):
        for col_idx in range(1, len(header) + 1):
            if col_idx != 2 + (1 if add_flags_column else 0):  # Кроме столбца "Команда"
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="right")
    
    # Заполняем данные о вопросах
    # Вычисляем смещение из-за добавления столбца с флагами
    flags_offset = 1 if add_flags_column else 0
    start_col = len(header) - 3 - len(valid_team_ids)
    
    for idx, question_tuple in enumerate(question_data, start=2):
        # Распаковываем данные о вопросе
        q_num = question_tuple[0]
        taken = question_tuple[1]
        rating = question_tuple[2]
        cat = question_tuple[3]
        
        # Записываем основные данные о вопросе
        ws.cell(row=idx, column=start_col).value = q_num
        ws.cell(row=idx, column=start_col + 1).value = taken
        ws.cell(row=idx, column=start_col + 2).value = rating
        ws.cell(row=idx, column=start_col + 3).value = cat
        
        # Записываем данные о взятиях выбранных команд
        for i, team_id in enumerate(valid_team_ids):
            team_col = start_col + 4 + i
            if len(question_tuple) > 4 + i:
                ws.cell(row=idx, column=team_col).value = question_tuple[4 + i]
        
        # Выравнивание по правому краю для данных о вопросах
        for col in range(start_col, start_col + 4 + len(valid_team_ids)):
            if col <= len(header):  # Проверка, чтобы не выйти за пределы
                ws.cell(row=idx, column=col).alignment = Alignment(horizontal="right")
    
    # Создаем имя файла с учетом флагов
    filename = f"{args.id}"
    if filter_flag_ids:
        filename += f"_flags_{'_'.join(map(str, filter_flag_ids))}"
    wb.save(f"{filename}.xlsx")
    print(f"Файл {filename}.xlsx успешно создан")

if __name__ == "__main__":
    main()
